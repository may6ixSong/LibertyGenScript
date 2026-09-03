"""
port_list_reader.py

Port List Excel 파일(.xls/.xlsx)을 읽어 대상 시트를 찾고, 헤더 행을 유동적으로
탐지한 뒤 필수 컬럼 값 존재 여부를 검사하고 port 개수를 세는 로직.
GUI에 의존하지 않는 순수 함수로 작성.

- .xlsx 는 openpyxl 로 읽음
- .xls (구 포맷) 는 openpyxl이 지원하지 않아 xlrd 로 읽음
  (둘 다 없다면 이 파일 최초 import 시가 아니라, 실제 해당 확장자 파일을
   읽으려는 시점에 ImportError가 발생함 - 그때 필요한 것만 설치하면 됨.
   xlrd가 없을 때는 "무엇을 설치하면 되는지"가 그대로 Step1 Details에 뜨도록
   _import_xlrd()에서 메시지를 바꿔 올린다)
- 허용 확장자 목록은 field_defs.PORT_LIST_FILE_EXTENSIONS 하나로 관리한다.

2026-08 성능 재설계 (캐싱 + 조기 종료 - "빡센 파일 규칙"):
  예전에는 이 파일의 거의 모든 공개 함수(read_port_list/read_port_list_rows/
  list_pins_by_port_type/list_all_pin_names 등)가 각자 독립적으로 워크북을 열고
  `sheet.iter_rows()`로 시트 "전체"(선언된 사용 범위 전부)를 파이썬 리스트로
  통째로 읽어들였다. 실제 엔지니어링 Excel은 서식(테두리/색 등)이 열/행 전체에
  걸쳐 적용된 경우가 흔해서 사용 범위가 실제 데이터보다 훨씬 크게(수십만 행) 잡히는
  일이 매우 흔하고, 그 경우 매 호출마다 그 큰 범위를 처음부터 다시 스캔하는 게
  Step1 Validate/Step2/Step3/Step4에서 "화면이 완전히 멈춘 것처럼" 보이는 주요
  원인이었다.

  이제 두 가지로 이 문제를 없앤다:
  1. **캐싱**: 파일 하나당 (mtime, size)를 key로 파싱 결과를 캐시해 둔다
     (_parse_port_list_cached). 파일이 안 바뀌었으면 이후의 모든 호출은 디스크를
     다시 읽지 않고 캐시된 결과만 재사용한다 - Step2/3/4에서 반복적으로 읽던 비용이
     세션당 1회로 줄어든다.
  2. **조기 종료(더 빡센 파일 규칙)**: 헤더 다음부터 데이터를 읽되, 완전히 빈 행이
     `_MAX_TRAILING_BLANK_ROWS`(500)개 연속으로 나오면 그 지점을 데이터의 끝으로
     간주하고 더 이상 읽지 않는다. 즉 **Port List의 실제 데이터 구간에 500행을
     넘는 완전 공백 gap이 있으면 안 된다**는 규칙을 강제한다 - 서식만 있고 값은
     없는 나머지 수십만 행을 절대 스캔하지 않게 되어, 파일의 "선언된 크기"가 아니라
     "실제 데이터 크기 + 여유분"에 비례하는 시간만 걸린다. 또한 읽는 열도 헤더에서
     실제로 인식된 컬럼까지만으로 제한한다(서식이 걸린 먼 오른쪽 열까지 읽지 않음).
"""

from __future__ import annotations

import re
from pathlib import Path

from step1_setup.field_defs import (
    PORT_LIST_FILE_EXTENSIONS, PORT_LIST_OPTIONAL_COLUMNS, PORT_LIST_REQUIRED_COLUMNS,
    PORT_TYPE_GND, PORT_TYPE_IO, PORT_TYPE_PWR,
)

# read_port_list_rows()가 각 행마다 담아주는 컬럼 전체 (필수 + 선택).
# 이 순서/집합에 없는 컬럼은 구조화 결과에 포함되지 않음 (필요해지면 여기에 추가).
_ALL_ROW_COLUMNS = PORT_LIST_REQUIRED_COLUMNS + PORT_LIST_OPTIONAL_COLUMNS

# 시트에 나타날 수 있는 헤더 문자열(대소문자/공백/기호 무시하고 매칭)을
# 표준 컬럼명으로 정규화하기 위한 매핑
_HEADER_ALIASES = {
    "port": "Port",
    "pinname": "Pin name",
    "bits": "Bits",
    "num": "Num",
    "io": "I/O",
    "volts": "Volts",
    "cap": "Cap",
    "map": "Map",
    "relatedpower": "Related Power",
    "relatedground": "Related ground",
    "relatedpin": "Related Pin",
    "timingreference": "Timing reference",
    "block": "Block",
    "maxtransition": "Max transition",
    "functionaldescription": "Functional Description",
    "description": "Description",
    "type": "Type",
    "defaultbinaryhex": "Default(binary/hex)",
    "digitalreg": "Digital Reg",
}

_MAX_HEADER_SCAN_ROWS = 5
_SHEET_NAME_HINT = "port list"

# 2026-08 추가 - 데이터 조기 종료 규칙: 헤더 이후 이 개수만큼 완전히 빈 행이 연속되면
# 그 지점을 데이터 끝으로 보고 더 이상 읽지 않는다. 실제 포트 개수가 아무리 많아도
# (수천~수만 개) 이 정도의 연속 gap은 나오지 않는다고 가정하는, 의도적으로 엄격한
# 규칙이다 - 이보다 큰 의도적인 공백 구간이 있는 Port List는 지원하지 않는다.
_MAX_TRAILING_BLANK_ROWS = 500


def _normalize(text: str) -> str:
    return "".join(ch for ch in text.lower() if ch.isalnum())


def _is_valid_int(value) -> bool:
    """
    Bits 컬럼 값 검사용. 정수만 허용(소수점 있는 값은 불허), 엑셀에서 숫자 셀은
    float(예: 4.0)로 들어올 수 있으므로 그 경우도 정수로 취급한다.
    """
    if value is None:
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value.is_integer()
    text = str(value).strip()
    if not text:
        return False
    try:
        return float(text).is_integer()
    except ValueError:
        return False


def _to_int(value) -> int:
    return int(float(value))


_VOLTS_NUMBER_PATTERN = re.compile(r"[-+]?\d*\.?\d+")


def parse_volts_value(value) -> float | None:
    """
    Port List 'Volts' 컬럼 값에서 숫자만 뽑아 float으로 반환한다. 단위가 붙어있는
    경우(예: "0.8V", "0.8 V")도 앞의 숫자 부분만 파싱한다. 숫자를 찾을 수 없으면
    None (예외를 던지지 않음 - block4의 pg_pin 작성에서 결측 처리로 이어짐).
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    match = _VOLTS_NUMBER_PATTERN.search(text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _import_xlrd():
    """
    .xls(구 엑셀 포맷)를 읽는 데 필요한 xlrd를 import한다. 없으면 무슨 패키지가
    필요한지 바로 알 수 있는 메시지로 바꿔서 올린다 (Step1 Validate 화면의 Details에
    그대로 표시된다).
    """
    try:
        import xlrd  # noqa: PLC0415 - 확장자가 .xls일 때만 필요한 선택적 의존성
    except ImportError as e:
        raise ImportError(
            "Reading an .xls (old Excel format) file requires the 'xlrd' package. "
            "Install it (pip install xlrd) or save the Port List as .xlsx."
        ) from e
    return xlrd


def _suffix_of(file_path: str) -> str:
    suffix = Path(file_path).suffix.lower()
    if suffix not in PORT_LIST_FILE_EXTENSIONS:
        raise ValueError(
            "Unsupported Port List file extension: %s (expected %s)"
            % (suffix or "(none)", " / ".join(PORT_LIST_FILE_EXTENSIONS))
        )
    return suffix


def _list_sheet_names(file_path: str) -> list[str]:
    if _suffix_of(file_path) == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()

    xlrd = _import_xlrd()
    wb = xlrd.open_workbook(file_path, on_demand=True)
    return list(wb.sheet_names())


def find_port_list_sheet_name(file_path: str) -> str | None:
    """시트명에 (대소문자 무관) 'port list'가 포함된 첫 시트명을 반환."""
    if not file_path or not Path(file_path).is_file():
        return None
    for name in _list_sheet_names(file_path):
        if _SHEET_NAME_HINT in name.lower():
            return name
    return None


def _find_header_row(rows: list[list]) -> tuple[int, dict[str, int]]:
    """
    상단 몇 개 행(_MAX_HEADER_SCAN_ROWS)을 스캔해서 알려진 컬럼명과 가장 많이
    일치하는 행을 헤더 행으로 판단 (1행일 수도, 2행일 수도 있으므로 유동 탐지).

    Returns:
        (header_row_index(1-based, 못 찾으면 0), {표준_컬럼명: col_index(1-based)})
    """
    best_row_idx = 0
    best_map: dict[str, int] = {}

    for row_idx, row in enumerate(rows[:_MAX_HEADER_SCAN_ROWS], start=1):
        header_map: dict[str, int] = {}
        for col_idx, value in enumerate(row, start=1):
            if value is None or str(value).strip() == "":
                continue
            canonical = _HEADER_ALIASES.get(_normalize(str(value)))
            if canonical:
                header_map[canonical] = col_idx
        if len(header_map) > len(best_map):
            best_map = header_map
            best_row_idx = row_idx

    return best_row_idx, best_map


def _is_blank_row(values) -> bool:
    return all(v is None or str(v).strip() == "" for v in values)


def _load_header_scan_rows(file_path: str, sheet_name: str) -> list[list]:
    """헤더 탐지용으로 처음 _MAX_HEADER_SCAN_ROWS행만 전체 열 폭으로 읽는다 (가벼움)."""
    if _suffix_of(file_path) == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = wb[sheet_name]
            return [
                list(row) for row in
                sheet.iter_rows(min_row=1, max_row=_MAX_HEADER_SCAN_ROWS, values_only=True)
            ]
        finally:
            wb.close()

    xlrd = _import_xlrd()
    wb = xlrd.open_workbook(file_path)
    sheet = wb.sheet_by_name(sheet_name)
    last_row = min(sheet.nrows, _MAX_HEADER_SCAN_ROWS)
    return [sheet.row_values(r) for r in range(last_row)]


def _load_bounded_data_rows(
    file_path: str, sheet_name: str, start_row: int, max_col: int,
) -> list[tuple[int, list]]:
    """
    start_row(1-based)부터, 열은 max_col까지만 읽는다. 완전히 빈 행이
    _MAX_TRAILING_BLANK_ROWS개 연속되면 그 지점에서 읽기를 멈춘다(2026-08 조기 종료
    규칙 - 모듈 docstring 참고). 빈 행 자체는 반환하지 않는다(호출부는 항상 빈 행을
    건너뛰므로 굳이 넘길 필요가 없다) - 다만 원본 Excel 행 번호는 그대로 유지해서
    반환하므로, 에러 메시지의 "Row N"은 항상 실제 엑셀 행 번호와 일치한다.

    Returns: [(row_number(1-based), values), ...] - 완전 공백이 아닌 행만.
    """
    if _suffix_of(file_path) == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = wb[sheet_name]
            rows: list[tuple[int, list]] = []
            blank_run = 0
            row_number = start_row - 1
            for values in sheet.iter_rows(min_row=start_row, max_col=max_col, values_only=True):
                row_number += 1
                if _is_blank_row(values):
                    blank_run += 1
                    if blank_run >= _MAX_TRAILING_BLANK_ROWS:
                        break
                    continue
                blank_run = 0
                rows.append((row_number, list(values)))
            return rows
        finally:
            wb.close()

    xlrd = _import_xlrd()
    wb = xlrd.open_workbook(file_path)
    sheet = wb.sheet_by_name(sheet_name)
    rows = []
    blank_run = 0
    for row_number in range(start_row, sheet.nrows + 1):
        values = sheet.row_values(row_number - 1, 0, max_col)
        if _is_blank_row(values):
            blank_run += 1
            if blank_run >= _MAX_TRAILING_BLANK_ROWS:
                break
            continue
        blank_run = 0
        rows.append((row_number, list(values)))
    return rows


# ---------------------------------------------------------------------------
# 캐시: 파일 하나당 (mtime, size)를 key로 파싱 결과를 저장한다. 이 프로세스가
# 살아있는 동안 같은 파일을 여러 번 읽어도(Step1 Validate -> Step2 콤보 -> Step3
# Check/Validate -> Step4 Generate) 디스크 접근은 파일이 바뀌지 않는 한 최초 1회뿐.
# ---------------------------------------------------------------------------
_PARSE_CACHE: dict[str, tuple] = {}


def _cache_key(file_path: str) -> tuple | None:
    try:
        st = Path(file_path).stat()
    except OSError:
        return None
    return (st.st_mtime_ns, st.st_size)


def clear_port_list_cache() -> None:
    """테스트/디버깅용 - 캐시를 강제로 비운다. 평소 운영 코드에서는 부를 필요 없음."""
    _PARSE_CACHE.clear()


def _parse_port_list_uncached(file_path: str) -> dict:
    sheet_name = find_port_list_sheet_name(file_path)
    if sheet_name is None:
        return {"sheet_name": None, "header_row": 0, "column_map": {}, "rows": []}

    header_scan_rows = _load_header_scan_rows(file_path, sheet_name)
    header_row, column_map = _find_header_row(header_scan_rows)

    rows: list[tuple[int, list]] = []
    if column_map:
        max_col = max(column_map.values())
        rows = _load_bounded_data_rows(file_path, sheet_name, header_row + 1, max_col)

    return {
        "sheet_name": sheet_name,
        "header_row": header_row,
        "column_map": column_map,
        "rows": rows,
    }


def _parse_port_list_cached(file_path: str) -> dict:
    """
    Port List 하나를 딱 한 번만 실제로 읽어서(헤더 탐지 + 조기 종료 규칙이 적용된
    데이터 행) 캐시해 두고, 이후 이 모듈의 모든 공개 함수가 이 캐시를 재사용한다.
    파일의 mtime/size가 바뀌면(수정/교체) 자동으로 다시 읽는다.

    Returns:
        {"sheet_name": str|None, "header_row": int, "column_map": {col:idx},
         "rows": [(row_number, values), ...]}
    """
    key = _cache_key(file_path)
    cached = _PARSE_CACHE.get(file_path)
    if cached is not None and key is not None and cached[0] == key:
        return cached[1]

    parsed = _parse_port_list_uncached(file_path)
    if key is not None:
        _PARSE_CACHE[file_path] = (key, parsed)
    return parsed


def read_port_list(file_path: str) -> dict:
    """
    Port List Excel 파일을 읽어 검증 결과를 반환.

    Returns:
        {
            "sheet_name": str | None,
            "header_row": int,
            "found_columns": [str, ...],
            "missing_required_columns": [str, ...],
            "port_count": int,
            "row_errors": [str, ...],   # "Row 12: missing value for Port" 형태
        }
    """
    parsed = _parse_port_list_cached(file_path)
    sheet_name = parsed["sheet_name"]
    if sheet_name is None:
        return {
            "sheet_name": None,
            "header_row": 0,
            "found_columns": [],
            "missing_required_columns": list(PORT_LIST_REQUIRED_COLUMNS),
            "port_count": 0,
            "row_errors": [],
        }

    column_map = parsed["column_map"]
    missing_required = [c for c in PORT_LIST_REQUIRED_COLUMNS if c not in column_map]

    row_errors: list[str] = []
    port_count = 0

    if not missing_required:
        for row_idx, row in parsed["rows"]:
            def _value_at(col_idx: int, _row=row):
                return _row[col_idx - 1] if col_idx - 1 < len(_row) else None

            # 조기 종료 규칙 적용 후에는 완전 공백 행이 결과에 없지만, 그래도
            # required 컬럼만 빈 "부분 공백" 행은 여전히 있을 수 있으니 방어적으로 유지.
            required_values = [_value_at(column_map[c]) for c in PORT_LIST_REQUIRED_COLUMNS]
            if all(v is None or str(v).strip() == "" for v in required_values):
                continue

            missing_in_row = [
                c for c in PORT_LIST_REQUIRED_COLUMNS
                if _value_at(column_map[c]) is None or str(_value_at(column_map[c])).strip() == ""
            ]
            if missing_in_row:
                row_errors.append(f"Row {row_idx}: missing value for {', '.join(missing_in_row)}")
                continue

            # 2026-08 확정: Bits 컬럼은 정수여야 함 (Step4 block3의 type_bus 생성이
            # 이 값을 그대로 bit_width/bit_from으로 사용하므로, 숫자가 아니면 여기서
            # 걸러내야 함).
            bits_value = _value_at(column_map["Bits"])
            if not _is_valid_int(bits_value):
                row_errors.append(f"Row {row_idx}: Bits value is not a valid integer: {bits_value!r}")
                continue

            port_count += 1

    return {
        "sheet_name": sheet_name,
        "header_row": parsed["header_row"],
        "found_columns": sorted(column_map.keys()),
        "missing_required_columns": missing_required,
        "port_count": port_count,
        "row_errors": row_errors,
    }


def list_pins_by_port_type(file_path: str, port_type: str) -> list[str]:
    """Port List에서 'Port' 컬럼 값이 port_type과 일치하는 행들의 'Pin name' 값 목록을 반환."""
    parsed = _parse_port_list_cached(file_path)
    column_map = parsed["column_map"]
    if "Port" not in column_map or "Pin name" not in column_map:
        return []

    port_col = column_map["Port"]
    pin_col = column_map["Pin name"]
    target = port_type.strip().upper()

    result = []
    for _row_idx, row in parsed["rows"]:
        port_val = row[port_col - 1] if port_col - 1 < len(row) else None
        pin_val = row[pin_col - 1] if pin_col - 1 < len(row) else None
        if port_val is None or pin_val is None:
            continue
        if str(port_val).strip().upper() == target:
            name = str(pin_val).strip()
            if name:
                result.append(name)
    return result


def read_port_list_rows(file_path: str) -> dict:
    """
    Port List Excel을 읽어 'Port' 컬럼 값(PORT/PWR/GND) 기준으로 행을 분류해서
    구조화된 딕셔너리 리스트로 반환한다. (2026-08 확정: PORT=I/O, PWR=전원, GND=접지)

    이후 단계(.pdt/.udc/pg_pin 조립 등)에서 이 결과를 그대로 재사용할 수 있도록
    GUI에 의존하지 않는 순수 함수로 작성. 값이 없는 컬럼은 빈 문자열로 채운다.

    Returns:
        {
            "sheet_name": str | None,
            "port_rows": [ {컬럼명: 값, ...}, ... ],   # Port == "PORT" (I/O)
            "pwr_rows":  [ {...}, ... ],                # Port == "PWR"
            "gnd_rows":  [ {...}, ... ],                # Port == "GND"
            "unclassified_rows": [ {...}, ... ],        # Port 값이 PORT/PWR/GND 중 아무것도 아닌 행
            "errors": [str, ...],
        }
    """
    result = {
        "sheet_name": None,
        "port_rows": [],
        "pwr_rows": [],
        "gnd_rows": [],
        "unclassified_rows": [],
        "errors": [],
    }

    parsed = _parse_port_list_cached(file_path)
    sheet_name = parsed["sheet_name"]
    result["sheet_name"] = sheet_name
    if sheet_name is None:
        result["errors"].append("No sheet found with a name containing 'Port list'.")
        return result

    column_map = parsed["column_map"]
    missing_required = [c for c in PORT_LIST_REQUIRED_COLUMNS if c not in column_map]
    if missing_required:
        result["errors"].append(
            "Missing required column(s): " + ", ".join(missing_required)
        )
        return result

    def _value_at(row: list, col_name: str):
        col_idx = column_map.get(col_name)
        if col_idx is None:
            return ""
        raw = row[col_idx - 1] if col_idx - 1 < len(row) else None
        return "" if raw is None else raw

    for row_idx, row in parsed["rows"]:
        required_values = [_value_at(row, c) for c in PORT_LIST_REQUIRED_COLUMNS]
        if all(v is None or str(v).strip() == "" for v in required_values):
            continue  # 부분 공백 행 방어(완전 공백 행은 이미 파싱 단계에서 제외됨)

        record = {col: _value_at(row, col) for col in _ALL_ROW_COLUMNS}
        record["_row"] = row_idx

        port_type = str(record.get("Port", "")).strip().upper()
        if port_type == PORT_TYPE_IO:
            result["port_rows"].append(record)
        elif port_type == PORT_TYPE_PWR:
            result["pwr_rows"].append(record)
        elif port_type == PORT_TYPE_GND:
            result["gnd_rows"].append(record)
        else:
            result["unclassified_rows"].append(record)
            result["errors"].append(
                f"Row {row_idx}: unrecognized Port value '{record.get('Port', '')}' "
                "(expected PORT / PWR / GND)"
            )

    return result


def list_port_bit_values(file_path: str) -> list[int]:
    """
    Port List에서 'Port' 컬럼 값이 PORT(I/O)인 행들의 'Bits' 값을 정수로 모아
    중복 제거 후 오름차순으로 정렬해 반환한다 (Step4 block3의 type_bus 생성용).

    Step1 Validate에서 이미 Bits가 정수인지 검사하지만, 방어적으로 여기서도
    정수로 변환 안 되는 값은 조용히 건너뛴다(예외를 던지지 않음).
    """
    result = read_port_list_rows(file_path)
    values: set[int] = set()
    for row in result["port_rows"]:
        bits = row.get("Bits", "")
        if _is_valid_int(bits):
            values.add(_to_int(bits))
    return sorted(values)


def list_power_ground_pins(file_path: str) -> dict:
    """
    Port List에서 Port=="PWR" / "GND"인 행들을 각각 파일에 나온 순서 그대로
    {"pin_name": str, "voltage_value": float | None, "related_pins": [str, ...]}
    리스트로 반환한다 (Step4 block4의 pg_pin, block5의
    {process_prefix}_pdt_pin(...) 생성용).

    - voltage_value는 'Volts' 컬럼에서 단위를 뺀 숫자만 파싱한 값이며, 파싱이 안
      되면 None(예외를 던지지 않고 결측 처리로 이어짐).
    - related_pins(2026-08 추가): Port List 전체(타입 무관)를 훑어서, 이 핀 이름을
      'Related Power'(PWR인 경우) 또는 'Related ground'(GND인 경우) 컬럼 값으로
      쓰고 있는 다른 행들의 Pin name을 파일에 나온 순서대로 모은 것. PWR pin은
      Related Power 컬럼만 보고, GND pin은 Related ground 컬럼만 본다(서로 교차해서
      보지 않음).

    Returns:
        {"pwr_pins": [{...}, ...], "gnd_pins": [{...}, ...]}
    """
    result = read_port_list_rows(file_path)
    all_rows = (
        result["port_rows"] + result["pwr_rows"] + result["gnd_rows"] + result["unclassified_rows"]
    )
    all_rows.sort(key=lambda r: r.get("_row", 0))

    def _to_pin_list(rows: list[dict], related_column: str) -> list[dict]:
        pins = []
        for row in rows:
            pin_name = str(row.get("Pin name", "")).strip()
            if not pin_name:
                continue
            related_pins = [
                str(other.get("Pin name", "")).strip()
                for other in all_rows
                if str(other.get(related_column, "")).strip() == pin_name
                and str(other.get("Pin name", "")).strip()
            ]
            pins.append({
                "pin_name": pin_name,
                "voltage_value": parse_volts_value(row.get("Volts")),
                "related_pins": related_pins,
            })
        return pins

    return {
        "pwr_pins": _to_pin_list(result["pwr_rows"], "Related Power"),
        "gnd_pins": _to_pin_list(result["gnd_rows"], "Related ground"),
    }


def list_port_pins_detailed(file_path: str) -> list[dict]:
    """
    Port List에서 Port=="PORT"인 행들을 파일에 나온 순서 그대로
    {"pin_name": str, "bits": int, "volts": float | None, "cap": float | None,
     "related_power": str, "related_ground": str, "related_pin": str, "io": str,
     "map": str, "type": str}
    리스트로 반환한다 (Step4 block5의 pin()/bus() 생성용).

    - bits는 Step1 Validate에서 이미 정수인지 검사하지만, 방어적으로 여기서도 정수로
      변환 안 되면 해당 행은 건너뛴다(예외를 던지지 않음).
    - volts는 단위(V 등)가 붙어있어도 숫자만 파싱한다. 파싱 안 되면 None.
    - cap은 숫자로 바로 파싱을 시도하고, 안 되면 None.
    - related_power/related_ground/related_pin/io/map/type은 빈 값이면 빈 문자열 그대로
      반환(결측 표시는 block5_writer.py에서 처리). map은 block5의 `is_analog` 판단에
      쓰인다(2026-08 추가, "Map" 컬럼 값이 대소문자 무시로 "analog"와 같을 때만). type은
      block5의 `{process_prefix}_pin_type` 값 결정에 쓰인다("Type" 컬럼 값이 대소문자
      무시로 "clock"과 같을 때만 "clock"으로 씀, 그 외에는 기존 data/data_bus 그대로).
    """
    result = read_port_list_rows(file_path)
    pins: list[dict] = []
    for row in result["port_rows"]:
        pin_name = str(row.get("Pin name", "")).strip()
        if not pin_name:
            continue
        bits_raw = row.get("Bits", "")
        if not _is_valid_int(bits_raw):
            continue
        cap_raw = row.get("Cap", "")
        try:
            cap_value = float(str(cap_raw).strip())
        except (TypeError, ValueError):
            cap_value = None

        pins.append({
            "pin_name": pin_name,
            "bits": _to_int(bits_raw),
            "volts": parse_volts_value(row.get("Volts")),
            "cap": cap_value,
            "related_power": str(row.get("Related Power", "")).strip(),
            "related_ground": str(row.get("Related ground", "")).strip(),
            "related_pin": str(row.get("Related Pin", "")).strip(),
            "io": str(row.get("I/O", "")).strip(),
            "map": str(row.get("Map", "")).strip(),
            "type": str(row.get("Type", "")).strip(),
        })
    return pins


def list_all_pin_names(file_path: str) -> list[str]:
    """Port List의 모든 행에서 'Pin name' 값 목록을 반환 (Port 타입 무관)."""
    parsed = _parse_port_list_cached(file_path)
    column_map = parsed["column_map"]
    if "Pin name" not in column_map:
        return []

    pin_col = column_map["Pin name"]
    result = []
    for _row_idx, row in parsed["rows"]:
        pin_val = row[pin_col - 1] if pin_col - 1 < len(row) else None
        if pin_val is not None:
            name = str(pin_val).strip()
            if name:
                result.append(name)
    return result


# ---------------------------------------------------------------------------
# 2026-08 추가 - DBS output pin bit 분할: 'Pin name'의 '[MSB:LSB]' 범위 표기를
# 다루는 공용 헬퍼. block5_writer.py(step4)와 pin_field_defs.py/settings_validator.py
# (step3)가 둘 다 이 모듈(step1)에서 가져다 쓴다 - step1은 그 위 step을 참조하지
# 않으므로 순환 import 없이 공유할 수 있는 유일한 위치다.
# ---------------------------------------------------------------------------
_BIT_RANGE_SUFFIX_RE = re.compile(r"\[(\d+):(\d+)\]\s*$")


def strip_bit_range_suffix(pin_name: str) -> str:
    """'BUS0[3:0]' -> 'BUS0'. 대괄호가 없으면 그대로 반환."""
    idx = pin_name.find("[")
    return pin_name if idx == -1 else pin_name[:idx]


def parse_bit_range(pin_name: str, bits: int) -> tuple[int, int]:
    """
    'BUS0[15:0]' -> (15, 0) (MSB, LSB). pin_name 끝에 '[MSB:LSB]'가 없으면, Port
    List가 0-based로 선언했다고 가정하고 (bits-1, 0)을 기본값으로 돌려준다.
    """
    match = _BIT_RANGE_SUFFIX_RE.search((pin_name or "").strip())
    if match:
        return int(match.group(1)), int(match.group(2))
    return bits - 1, 0


def list_all_pin_bit_info(file_path: str) -> dict[str, dict]:
    """
    Port List 전체 행(Port 컬럼 값이 PORT/PWR/GND/그 외 무엇이든 전부)에서, 대괄호
    범위를 뺀 base pin name -> {"bits": int, "msb": int, "lsb": int, "full_name": str}
    매핑을 만든다.

    2026-08 DBS output pin bit 분할 추가: Step3 "Check DBS Output Pins"로 인식한 DBS
    output pin마다 Port List의 'Related Pin' 컬럼이 가리키는 pin은 Port=="PORT"가
    아닐 수도 있으므로(보장되지 않음), Port=="PORT"만 보는 list_port_pins_detailed와
    달리 이 함수는 전체 행을 본다 - 그 pin의 Bits 값을 찾아 DBS output pin과 같은
    quotient(그룹 개수)로 나눈 related_bus_pins 범위를 계산하는 데 쓰인다
    (block5_writer.py, settings_validator.py, settings_view.py 공용).

    Bits가 정수로 읽히지 않는 행은 건너뛴다. 같은 base name이 여러 행에 나오면
    처음 나온 행을 쓴다.
    """
    parsed = _parse_port_list_cached(file_path)
    column_map = parsed["column_map"]
    if "Pin name" not in column_map or "Bits" not in column_map:
        return {}

    pin_col = column_map["Pin name"]
    bits_col = column_map["Bits"]
    result: dict[str, dict] = {}
    for _row_idx, row in parsed["rows"]:
        pin_val = row[pin_col - 1] if pin_col - 1 < len(row) else None
        bits_val = row[bits_col - 1] if bits_col - 1 < len(row) else None
        if pin_val is None or not _is_valid_int(bits_val):
            continue
        name = str(pin_val).strip()
        if not name:
            continue
        bits = _to_int(bits_val)
        msb, lsb = parse_bit_range(name, bits)
        base = strip_bit_range_suffix(name)
        if base not in result:
            result[base] = {"bits": bits, "msb": msb, "lsb": lsb, "full_name": name}
    return result
